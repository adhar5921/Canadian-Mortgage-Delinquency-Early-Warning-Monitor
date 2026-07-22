"""Convert the CMHC workbook to tidy CSV and a queryable SQLite database."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_FILE = ROOT / "data" / "raw" / "cmhc_mortgage_delinquency_2012Q3_2025Q4.xlsx"
PROCESSED_FILE = ROOT / "data" / "processed" / "mortgage_delinquency_long.csv"
DATABASE_FILE = ROOT / "mortgage_risk.db"
SCHEMA_FILE = ROOT / "sql" / "00_schema.sql"


def extract_records() -> pd.DataFrame:
    """Read the semi-structured workbook and return one row per geography-quarter."""
    workbook = load_workbook(RAW_FILE, data_only=True, read_only=True)
    sheet = workbook["Mortgage delinquency rate"]

    quarters = [sheet.cell(5, column).value for column in range(3, 57)]
    geography_type = "country"
    records: list[dict[str, object]] = []

    for row in range(6, 53):
        geography = sheet.cell(row, 2).value

        if geography == "Provinces":
            geography_type = "province"
            continue
        if geography == "Selected Metropolitan Areas":
            geography_type = "cma"
            continue
        if not geography:
            continue

        for column, quarter in enumerate(quarters, start=3):
            value = sheet.cell(row, column).value
            if value in (None, ""):
                continue

            quarter_text = str(quarter)
            records.append(
                {
                    "geography": geography,
                    "geography_type": geography_type,
                    "quarter": quarter_text,
                    "year": int(quarter_text[:4]),
                    "quarter_number": int(quarter_text[-1]),
                    "quarter_index": int(quarter_text[:4]) * 4 + int(quarter_text[-1]),
                    "delinquency_rate_pct": float(value),
                }
            )

    data = pd.DataFrame.from_records(records)
    data = data.sort_values(["geography_type", "geography", "quarter_index"]).reset_index(drop=True)
    return data


def validate(data: pd.DataFrame) -> None:
    """Fail loudly if source shape or key assumptions change."""
    expected_columns = {
        "geography",
        "geography_type",
        "quarter",
        "year",
        "quarter_number",
        "quarter_index",
        "delinquency_rate_pct",
    }
    assert set(data.columns) == expected_columns, "Unexpected output columns"
    assert len(data) == 2_430, "Expected 45 geographies x 54 quarters"
    assert data["geography"].nunique() == 45, "Unexpected geography count"
    assert not data.duplicated(["geography", "quarter"]).any(), "Duplicate geography-quarter key"
    assert not data.isna().any().any(), "Missing values in tidy data"
    assert data["delinquency_rate_pct"].between(0, 100).all(), "Rate outside 0%-100%"
    assert set(data["quarter_number"].unique()) == {1, 2, 3, 4}, "Invalid quarter number"


def save_outputs(data: pd.DataFrame) -> None:
    PROCESSED_FILE.parent.mkdir(parents=True, exist_ok=True)
    data.to_csv(PROCESSED_FILE, index=False)

    with sqlite3.connect(DATABASE_FILE) as connection:
        connection.executescript(SCHEMA_FILE.read_text(encoding="utf-8"))
        data.to_sql("mortgage_delinquency", connection, if_exists="append", index=False)


def main() -> None:
    data = extract_records()
    validate(data)
    save_outputs(data)
    print(
        f"Prepared {len(data):,} rows across {data['geography'].nunique()} geographies "
        f"and saved {PROCESSED_FILE.relative_to(ROOT)}."
    )


if __name__ == "__main__":
    main()
